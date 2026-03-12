from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
import logging
from pathlib import Path
from threading import Lock
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


logger = logging.getLogger("excel_backup")


@dataclass(frozen=True)
class ExcelBackupConfig:
    file_path: Path
    sheet_name: str = "users"


class ExcelBackupClient:
    def __init__(self, config: ExcelBackupConfig) -> None:
        self._config = config
        self._lock = Lock()

    @classmethod
    def with_default_path(cls) -> "ExcelBackupClient":
        project_root = Path(__file__).resolve().parent.parent
        return cls(ExcelBackupConfig(file_path=project_root / "users_backup.xlsx"))

    @property
    def file_path(self) -> Path:
        return self._config.file_path

    @staticmethod
    def _normalize_value(value: Any) -> Any:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.isoformat(sep=" ", timespec="seconds")
        if isinstance(value, date):
            return value.isoformat()
        return value

    @staticmethod
    def _ensure_header(worksheet: Worksheet, column_order: list[str]) -> None:
        has_header = any(worksheet.cell(row=1, column=index + 1).value for index in range(len(column_order)))
        if has_header:
            return

        for index, column_name in enumerate(column_order, start=1):
            worksheet.cell(row=1, column=index, value=column_name)

    def _load_or_create_sheet(self, column_order: list[str]) -> tuple[Workbook, Worksheet]:
        if self._config.file_path.exists():
            workbook = load_workbook(self._config.file_path)
            if self._config.sheet_name in workbook.sheetnames:
                worksheet = workbook[self._config.sheet_name]
                logger.debug(
                    "Opened existing Excel file path=%s sheet=%s",
                    self._config.file_path.name,
                    self._config.sheet_name,
                )
            else:
                worksheet = workbook.create_sheet(self._config.sheet_name)
                logger.info(
                    "Created new sheet in Excel file path=%s sheet=%s",
                    self._config.file_path.name,
                    self._config.sheet_name,
                )
        else:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = self._config.sheet_name
            logger.info("Created new Excel file path=%s", self._config.file_path.name)

        self._ensure_header(worksheet, column_order)
        return workbook, worksheet

    def append_row(self, row_data: dict[str, Any], column_order: list[str]) -> None:
        if not column_order:
            raise ValueError("Column order for Excel backup must not be empty.")

        values = [self._normalize_value(row_data.get(column_name)) for column_name in column_order]

        with self._lock:
            workbook, worksheet = self._load_or_create_sheet(column_order)
            worksheet.append(values)
            workbook.save(self._config.file_path)
            workbook.close()
            logger.info("Appended row to Excel path=%s", self._config.file_path.name)

    def replace_with_rows(self, rows: list[dict[str, Any]], column_order: list[str]) -> None:
        if not column_order:
            raise ValueError("Column order for Excel backup must not be empty.")

        with self._lock:
            workbook, worksheet = self._load_or_create_sheet(column_order)
            if worksheet.max_row > 0:
                worksheet.delete_rows(1, worksheet.max_row)

            for index, column_name in enumerate(column_order, start=1):
                worksheet.cell(row=1, column=index, value=column_name)

            for row in rows:
                values = [self._normalize_value(row.get(column_name)) for column_name in column_order]
                worksheet.append(values)

            workbook.save(self._config.file_path)
            workbook.close()
            logger.info("Rebuilt Excel backup path=%s rows=%s", self._config.file_path.name, len(rows))
